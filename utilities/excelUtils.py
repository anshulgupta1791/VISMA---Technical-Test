import openpyxl

def get_row_count(path, sheetNumber):
    wb = openpyxl.load_workbook(path)
    sheet = wb[sheetNumber]
    return (sheet.max_row)

def get_column_count(path, sheetNumber):
    wb = openpyxl.load_workbook(path)
    sheet = wb[sheetNumber]
    return (sheet.max_column)

def read_data(path, sheetNumber, rowNumber, columnNumber):
    wb = openpyxl.load_workbook(path)
    sheet = wb[sheetNumber]
    return sheet.cell(row=rowNumber, column=columnNumber).value

def write_data(path, sheetNumber, rowNumber, columnNumber, data):
    wb = openpyxl.load_workbook(path)
    sheet = wb[sheetNumber]
    sheet.cell(row=rowNumber, column=columnNumber).value = data
    wb.save(path)
